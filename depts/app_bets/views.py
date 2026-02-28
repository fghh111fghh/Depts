"""
Модуль анализа футбольных матчей с использованием нескольких статистических методов.

Класс AnalyzeView предоставляет:
1. GET: отображение формы для ввода данных
2. POST: обработка и анализ введенных данных
3. Парсинг текстовых данных с коэффициентами и названиями команд
4. Анализ матчей методами: Пуассона, "близнецов", паттернов форм, личных встреч
5. Векторный синтез для формирования итогового вердикта
6. Сохранение алиасов команд для улучшения распознавания

Основные сохраняемые структуры данных в контексте шаблона:
- results: список словарей с результатами анализа каждого матча
- unknown_teams: множество команд, которые не удалось распознать
- raw_text: оригинальный текст, введенный пользователем
- all_teams: QuerySet всех команд для выпадающего списка
"""
import csv
import os
import pickle
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Optional, Tuple
import openpyxl
import pandas as pd
from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.core.cache import cache
from django.db import transaction
from django.db.models import F, Q, Sum
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.utils.timezone import make_aware, get_current_timezone, now
from django.views import View
import re
import math
import unicodedata
from django.views.decorators.http import require_POST
from django.views.generic import TemplateView, CreateView, ListView
from openpyxl.styles import Font, PatternFill, Alignment
from dal import autocomplete
from app_bets.constants import Outcome, ParsingConstants, AnalysisConstants, Messages
from app_bets.forms import BetForm
from app_bets.models import Team, TeamAlias, Season, Match, League, Bet,Bank


class AnalyzeView(ListView):
    """
    Отображает список проанализированных матчей на основе загруженного текста.
    """
    template_name = 'app_bets/bets_main.html'
    context_object_name = 'results'
    paginate_by = 20

    def get_queryset(self):
        """
        Возвращает результаты из сессии.
        """
        results = self.request.session.get('results', [])
        current_sort = self.get_current_sort()
        results = self.sort_results(results, current_sort)
        return results

    def get_context_data(self, **kwargs):
        """
        Добавляет дополнительные данные в контекст шаблона.
        """
        context = super().get_context_data(**kwargs)
        context.update({
            'raw_text': self.request.session.get('raw_text', ''),
            'unknown_teams': sorted(self.request.session.get('unknown_teams', [])),
            'all_teams': Team.objects.all().order_by('name'),
            'current_sort': self.get_current_sort(),
        })
        return context

    def get_current_sort(self) -> str:
        """
        Возвращает текущий параметр сортировки.
        """
        return self.request.GET.get('sort') or self.request.session.get('current_sort', 'default')

    def sort_results(self, results: List[Dict], sort_param: str) -> List[Dict]:
        """
        Сортирует результаты в соответствии с параметром сортировки.
        """
        if not results:
            return results

        if sort_param == 'default':
            results[:] = [dict(r) for r in self.request.session.get('original_results', [])]
        elif sort_param == 'btts_desc':
            results.sort(key=lambda x: x['poisson_btts']['yes'], reverse=True)
        elif sort_param == 'over25_desc':
            results.sort(key=lambda x: x['poisson_over25']['yes'], reverse=True)
        elif sort_param == 'twins_p1_desc':
            results.sort(
                key=lambda x: max(
                    x.get('twins_data', {}).get('p1', 0) if x.get('twins_data') else 0,
                    x.get('twins_data', {}).get('p2', 0) if x.get('twins_data') else 0
                ),
                reverse=True
            )
        elif sort_param == 'pattern_p1_desc':
            results.sort(
                key=lambda x: max(
                    x.get('pattern_data', {}).get('p1', 0) if x.get('pattern_data') else 0,
                    x.get('pattern_data', {}).get('p2', 0) if x.get('pattern_data') else 0
                ),
                reverse=True
            )

        return results

    def post(self, request, *args, **kwargs):
        """
        Обрабатывает POST-запрос: парсинг текста и анализ матчей.
        """
        current_sort = request.POST.get('sort') or request.GET.get('sort') or request.session.get('current_sort',
                                                                                                  'default')
        request.session['current_sort'] = current_sort
        raw_text = request.POST.get('matches_text', '')

        if 'create_alias' in request.POST:
            self._handle_alias_creation(request)

            if raw_text.strip():
                results, unknown_teams = self._analyze_matches(request, raw_text)
                request.session['results'] = results
                request.session['raw_text'] = raw_text
                request.session['unknown_teams'] = list(unknown_teams)

                if results:
                    request.session['original_results'] = [dict(r) for r in results]

            query_params = request.GET.urlencode()
            if query_params:
                return redirect(f"{request.path}?{query_params}")
            return redirect(request.path)

        if not raw_text.strip():
            return self._render_empty_response(request, raw_text, current_sort)

        results, unknown_teams = self._analyze_matches(request, raw_text)

        # СОХРАНЯЕМ ВСЕ ДАННЫЕ В СЕССИЮ
        request.session['results'] = results
        request.session['raw_text'] = raw_text
        request.session['unknown_teams'] = list(unknown_teams)

        if results:
            request.session['original_results'] = [dict(r) for r in results]

        # РЕДИРЕКТ НА GET С ПАРАМЕТРАМИ СОРТИРОВКИ
        return redirect(f"{request.path}?sort={current_sort}")

    def _handle_alias_creation(self, request):
        """
        Создает новый алиас для команды.
        """
        alias_raw = request.POST.get('alias_name', '').strip()
        team_id = request.POST.get('team_id', '').strip()

        if not alias_raw:
            messages.error(request, "Название команды не может быть пустым")
            return False

        if not team_id:
            messages.error(request, "Не выбрана команда из списка")
            return False

        try:
            team = Team.objects.get(id=team_id)
            clean_name = " ".join(alias_raw.split()).lower()

            existing_alias = TeamAlias.objects.filter(name=clean_name).first()
            if existing_alias:
                if existing_alias.team.id == team.id:
                    messages.info(request, f"Привязка для '{alias_raw}' уже существует")
                    return True
                else:
                    messages.warning(
                        request,
                        f"Название '{alias_raw}' уже привязано к команде '{existing_alias.team.name}'. "
                        f"Сначала удалите старую привязку в админке."
                    )
                    return False

            with transaction.atomic():
                TeamAlias.objects.create(
                    name=clean_name,
                    team=team
                )

            cache.delete('match_analysis_full_data')

            messages.success(
                request,
                f"Команда '{alias_raw}' успешно привязана к '{team.name}'"
            )
            return True

        except Team.DoesNotExist:
            messages.error(request, f"Команда с ID {team_id} не найдена в базе")
            return False

        except Exception as e:
            messages.error(request, f"Ошибка при создании привязки: {e}")
            return False

    def _render_empty_response(self, request, raw_text: str, current_sort: str):
        """
        Возвращает пустой ответ, когда нет данных для анализа.
        """
        return render(request, self.template_name, {
            'results': [],
            'raw_text': raw_text,
            'unknown_teams': [],
            'all_teams': Team.objects.all().order_by('name'),
            'current_sort': current_sort,
        })

    def _load_cached_data(self):
        """
        Загружает данные с кэшированием в памяти.
        """
        cache_key = 'match_analysis_full_data'

        cached_data = cache.get(cache_key)
        if cached_data is not None:
            return cached_data

        all_matches = list(Match.objects.filter(
            home_score_reg__isnull=False
        ).select_related(
            'home_team', 'away_team', 'league', 'season'
        ).order_by('date'))

        matches_by_league = {}
        for match in all_matches:
            if match.league_id not in matches_by_league:
                matches_by_league[match.league_id] = []
            matches_by_league[match.league_id].append(match)

        all_teams = {team.id: team for team in Team.objects.all()}

        all_aliases = {}
        for alias in TeamAlias.objects.all().select_related('team'):
            all_aliases[alias.name] = alias.team

        all_leagues = {league.id: league for league in League.objects.all()}

        data = (all_matches, matches_by_league, all_teams, all_aliases, all_leagues)
        cache.set(cache_key, data, 86400)

        return data

    def _analyze_matches(self, request, raw_text: str) -> Tuple[List[Dict], set]:
        """
        Основной метод анализа матчей.
        """
        results = []
        unknown_teams = set()

        season = Season.objects.filter(is_current=True).first() or Season.objects.order_by('-start_date').first()

        lines = [l.strip() for l in raw_text.split('\n') if l.strip()]
        if not lines:
            return results, unknown_teams

        all_matches, matches_by_league, all_teams_dict, all_aliases, all_leagues = self._load_cached_data()

        skip_to = -1
        for i, line in enumerate(lines):
            if i <= skip_to:
                continue

            if re.match(ParsingConstants.ODDS_REGEX, line):
                try:
                    match_data = self._parse_match_data(lines, i)
                    if not match_data:
                        continue

                    h_odd, d_odd, a_odd, skip_to, names = match_data

                    if len(names) != 2:
                        continue

                    away_raw, home_raw = names[0], names[1]

                    home_team, away_team = self._find_teams(
                        home_raw, away_raw, all_teams_dict, all_aliases
                    )

                    if home_team and away_team:
                        result = self._analyze_single_match(
                            home_team, away_team, season, all_matches, matches_by_league,
                            all_leagues, h_odd, d_odd, a_odd
                        )
                        if result:
                            results.append(result)
                    else:
                        if not home_team:
                            unknown_teams.add(home_raw.strip())
                        if not away_team:
                            unknown_teams.add(away_raw.strip())

                except (IndexError, ValueError, Exception):
                    continue

        return results, unknown_teams

    def _parse_match_data(self, lines: List[str], index: int) -> Optional[Tuple]:
        """
        Парсит коэффициенты и названия команд из строк.
        """
        try:
            h_odd = Decimal(lines[index].replace(',', '.')).quantize(Decimal(Messages.DECIMAL_FORMAT))
            d_odd = Decimal(lines[index + 1].replace(',', '.')).quantize(Decimal(Messages.DECIMAL_FORMAT))
            a_odd = Decimal(lines[index + 2].replace(',', '.')).quantize(Decimal(Messages.DECIMAL_FORMAT))
            skip_to = index + 2

            names = self._extract_team_names(lines, index)

            return h_odd, d_odd, a_odd, skip_to, names
        except Exception:
            return None

    def _extract_team_names(self, lines: List[str], odds_index: int) -> List[str]:
        """
        Извлекает названия команд из строк перед коэффициентами.
        """
        names = []
        search_depth = min(ParsingConstants.MAX_SEARCH_DEPTH, odds_index)

        for j in range(odds_index - 1, odds_index - search_depth - 1, -1):
            if j < 0:
                break

            row = lines[j].strip()

            if (not row or row == '-' or
                    re.match(ParsingConstants.TIME_REGEX, row) or
                    row.lower() in ParsingConstants.SKIP_KEYWORDS or
                    any(keyword in row.lower() for keyword in ParsingConstants.LEAGUE_KEYWORDS)):
                continue

            clean_name = self.clean_team_name(row)
            if (clean_name and
                    len(clean_name) >= AnalysisConstants.MIN_TEAM_NAME_LENGTH and
                    not re.match(ParsingConstants.DIGITS_ONLY_REGEX, clean_name) and
                    clean_name not in ParsingConstants.BLACKLIST):

                names.append(row)
                if len(names) == 2:
                    break

        return names

    @staticmethod
    def clean_team_name(name: str) -> str:
        """
        Очищает название команды для поиска.
        """
        if not name:
            return ""
        try:
            name = unicodedata.normalize('NFKC', str(name))
            name = re.sub(ParsingConstants.TIME_REGEX, '', name)
            name = re.sub(r'[^\w\s\d\-\'\(\)]', ' ', name)
            name = re.sub(r'^\d+\s+|\s+\d+$', '', name)
            name = re.sub(r'[\-\–\—]+', ' ', name)
            name = ' '.join(name.split())
            return name.strip().lower()
        except Exception:
            return str(name).strip().lower() if name else ""

    def _find_teams(self, home_raw: str, away_raw: str, all_teams: Dict, all_aliases: Dict) -> Tuple[
        Optional[object], Optional[object]]:
        """
        Находит команды по их названиям.
        """
        clean_home = self.clean_team_name(home_raw)
        clean_away = self.clean_team_name(away_raw)

        home_team = self._find_team(clean_home, all_teams, all_aliases)
        away_team = self._find_team(clean_away, all_teams, all_aliases)

        return home_team, away_team

    def _find_team(self, clean_name: str, all_teams: Dict, all_aliases: Dict) -> Optional[object]:
        """
        Находит команду по очищенному названию.
        """
        if clean_name in all_aliases:
            return all_aliases[clean_name]

        for team in all_teams.values():
            if team.name.lower() == clean_name:
                return team

        clean_name_no_brackets = re.sub(r'\s*\([^)]*\)', '', clean_name).strip()
        if clean_name_no_brackets != clean_name:
            if clean_name_no_brackets in all_aliases:
                return all_aliases[clean_name_no_brackets]

            for team in all_teams.values():
                team_name_no_brackets = re.sub(r'\s*\([^)]*\)', '', team.name.lower()).strip()
                if team_name_no_brackets == clean_name_no_brackets:
                    return team

        return None

    def _analyze_single_match(self, home_team, away_team, season, all_matches, matches_by_league,
                              all_leagues, h_odd, d_odd, a_odd) -> Optional[Dict]:
        """
        Анализирует один матч.
        """
        league = self._determine_league(
            home_team, away_team, season, all_matches, matches_by_league, all_leagues
        )

        if not league:
            return None

        league_matches = matches_by_league.get(league.id, [])

        pattern_data, curr_h_form, curr_a_form = self._analyze_pattern(
            home_team, away_team, season, league_matches
        )

        poisson_results, p_data = self._calculate_poisson(
            home_team, away_team, league, season, h_odd
        )

        twins_data, t_count = self._find_twins_matches(
            league_matches, h_odd, a_odd
        )

        h2h_list = self._get_h2h_matches(home_team, away_team)

        m_obj = Match(
            home_team=home_team,
            away_team=away_team,
            league=league,
            season=season,
            odds_home=h_odd
        )
        historical_total_insight = m_obj.get_historical_total_insight()

        return {
            'match': f"{home_team.name} - {away_team.name}",
            'league': league.name if league else "Unknown",
            'poisson_l': f"{p_data['home_lambda']:.2f} : {p_data['away_lambda']:.2f}",
            'poisson_top': poisson_results['top_scores'],
            'poisson_btts': {
                'yes': poisson_results['btts_yes'],
                'no': poisson_results['btts_no']
            },
            'poisson_over25': {
                'yes': poisson_results['over25_yes'],
                'no': poisson_results['over25_no']
            },
            'twins_count': t_count,
            'twins_data': twins_data,
            'pattern_data': pattern_data,
            'current_h_form': curr_h_form,
            'current_a_form': curr_a_form,
            'h2h_list': h2h_list,
            'h2h_total': len(h2h_list),
            'odds': (float(h_odd), float(d_odd), float(a_odd)),
            'historical_total': historical_total_insight.get('synthetic'),
        }

    def _determine_league(self, home_team, away_team, season, all_matches, matches_by_league, all_leagues):
        """
        Определяет лигу для пары команд.
        """
        league = None
        current_season_matches = [m for m in all_matches if m.season_id == season.id]

        for match in current_season_matches:
            if ((match.home_team_id == home_team.id and match.away_team_id == away_team.id) or
                (match.home_team_id == away_team.id and match.away_team_id == home_team.id)) and match.league:
                league = match.league
                break

        if not league:
            for match in current_season_matches:
                if match.home_team_id == home_team.id and match.league:
                    league = match.league
                    break

        if not league:
            for match in current_season_matches:
                if match.away_team_id == away_team.id and match.league:
                    league = match.league
                    break

        if not league:
            for match in all_matches:
                if ((match.home_team_id == home_team.id and match.away_team_id == away_team.id) or
                    (match.home_team_id == away_team.id and match.away_team_id == home_team.id)) and match.league:
                    league = match.league
                    break

        if not league:
            for match in all_matches:
                if match.home_team_id == home_team.id and match.league:
                    league = match.league
                    break

        if not league:
            for match in all_matches:
                if match.away_team_id == away_team.id and match.league:
                    league = match.league
                    break

        return league

    def _analyze_pattern(self, home_team, away_team, season, league_matches):
        """
        Анализирует исторические паттерны формы команд.
        """
        current_season_matches = [m for m in league_matches if m.season_id == season.id]
        sorted_current_season_matches = sorted(current_season_matches, key=lambda x: x.date)

        team_history_current = {}

        for m in sorted_current_season_matches:
            h_id, a_id = m.home_team_id, m.away_team_id

            if m.home_score_reg == m.away_score_reg:
                res_h = Outcome.DRAW
                res_a = Outcome.DRAW
            elif m.home_score_reg > m.away_score_reg:
                res_h = Outcome.WIN
                res_a = Outcome.LOSE
            else:
                res_h = Outcome.LOSE
                res_a = Outcome.WIN

            team_history_current.setdefault(h_id, []).append(res_h)
            team_history_current.setdefault(a_id, []).append(res_a)

        curr_h_form = "".join(team_history_current.get(home_team.id, []))[-AnalysisConstants.PATTERN_FORM_LENGTH:]
        curr_a_form = "".join(team_history_current.get(away_team.id, []))[-AnalysisConstants.PATTERN_FORM_LENGTH:]

        all_league_matches_sorted = sorted(league_matches, key=lambda x: x.date)
        team_history_all = {}
        match_patterns_all = {}

        for m in all_league_matches_sorted:
            h_id, a_id = m.home_team_id, m.away_team_id

            h_f = "".join(team_history_all.get(h_id, []))[-AnalysisConstants.PATTERN_FORM_LENGTH:]
            a_f = "".join(team_history_all.get(a_id, []))[-AnalysisConstants.PATTERN_FORM_LENGTH:]

            if len(h_f) == AnalysisConstants.PATTERN_FORM_LENGTH and len(a_f) == AnalysisConstants.PATTERN_FORM_LENGTH:
                match_patterns_all[m.id] = (h_f, a_f)

            if m.home_score_reg == m.away_score_reg:
                res_h = Outcome.DRAW
                res_a = Outcome.DRAW
            elif m.home_score_reg > m.away_score_reg:
                res_h = Outcome.WIN
                res_a = Outcome.LOSE
            else:
                res_h = Outcome.LOSE
                res_a = Outcome.WIN

            team_history_all.setdefault(h_id, []).append(res_h)
            team_history_all.setdefault(a_id, []).append(res_a)

        pattern_data = None
        p_hw, p_dw, p_aw, p_count = 0, 0, 0, 0

        if len(curr_h_form) == AnalysisConstants.PATTERN_FORM_LENGTH and len(
                curr_a_form) == AnalysisConstants.PATTERN_FORM_LENGTH:
            for m in league_matches:
                if match_patterns_all.get(m.id) == (curr_h_form, curr_a_form):
                    p_count += 1
                    if m.home_score_reg > m.away_score_reg:
                        p_hw += 1
                    elif m.home_score_reg == m.away_score_reg:
                        p_dw += 1
                    else:
                        p_aw += 1

            if p_count > 0:
                p1_pct = round(p_hw / p_count * 100)
                x_pct = round(p_dw / p_count * 100)
                p2_pct = round(p_aw / p_count * 100)

                total_pct = p1_pct + x_pct + p2_pct
                if total_pct != 100:
                    diff = 100 - total_pct
                    max_val = max(p1_pct, x_pct, p2_pct)
                    if p1_pct == max_val:
                        p1_pct += diff
                    elif x_pct == max_val:
                        x_pct += diff
                    else:
                        p2_pct += diff

                pattern_data = {
                    'pattern': f"{curr_h_form} - {curr_a_form}",
                    'count': p_count,
                    'p1': p1_pct,
                    'x': x_pct,
                    'p2': p2_pct
                }

        return pattern_data, curr_h_form, curr_a_form

    def _calculate_poisson(self, home_team, away_team, league, season, h_odd):
        """
        Рассчитывает вероятности по Пуассону.
        """
        m_obj = Match(
            home_team=home_team,
            away_team=away_team,
            league=league,
            season=season,
            odds_home=h_odd
        )
        p_data = m_obj.calculate_poisson_lambda(now())
        poisson_results = self.get_poisson_probs(p_data['home_lambda'], p_data['away_lambda'])

        return poisson_results, p_data

    def _find_twins_matches(self, league_matches, h_odd, a_odd):
        """
        Находит матчи-близнецы с похожими коэффициентами.
        """
        tol = AnalysisConstants.TWINS_TOLERANCE_SMALL
        twins_matches = []

        for m in league_matches:
            if m.odds_home and m.odds_away:
                h_diff = abs(float(m.odds_home) - float(h_odd))
                a_diff = abs(float(m.odds_away) - float(a_odd))
                if h_diff <= tol and a_diff <= tol:
                    twins_matches.append(m)

        if not twins_matches:
            tol = AnalysisConstants.TWINS_TOLERANCE_LARGE
            for m in league_matches:
                if m.odds_home and m.odds_away:
                    h_diff = abs(float(m.odds_home) - float(h_odd))
                    a_diff = abs(float(m.odds_away) - float(a_odd))
                    if h_diff <= tol and a_diff <= tol:
                        twins_matches.append(m)

        t_count = len(twins_matches)
        twins_data = None

        if t_count > 0:
            hw_t = sum(1 for m in twins_matches if m.home_score_reg > m.away_score_reg)
            dw_t = sum(1 for m in twins_matches if m.home_score_reg == m.away_score_reg)
            aw_t = sum(1 for m in twins_matches if m.home_score_reg < m.away_score_reg)

            total_with_results = hw_t + dw_t + aw_t

            if total_with_results > 0:
                p1_pct = round(hw_t / total_with_results * 100)
                x_pct = round(dw_t / total_with_results * 100)
                p2_pct = round(aw_t / total_with_results * 100)

                total_pct = p1_pct + x_pct + p2_pct
                if total_pct != 100:
                    diff = 100 - total_pct
                    max_val = max(p1_pct, x_pct, p2_pct)
                    if p1_pct == max_val:
                        p1_pct += diff
                    elif x_pct == max_val:
                        x_pct += diff
                    else:
                        p2_pct += diff

                twins_data = {
                    'count': t_count,
                    'p1': p1_pct,
                    'x': x_pct,
                    'p2': p2_pct
                }

        return twins_data, t_count

    def _get_h2h_matches(self, home_team, away_team):
        """
        Получает историю личных встреч команд.
        """
        h2h_queryset = Match.objects.filter(
            home_team=home_team,
            away_team=away_team
        ).select_related(
            'home_team', 'away_team'
        ).order_by('-date')[:10]

        h2h_list = []
        for m in h2h_queryset:
            h2h_list.append({
                'date': m.date.strftime(Messages.DATE_FORMAT),
                'score': f"{m.home_score_reg}:{m.away_score_reg}"
            })

        return h2h_list

    @staticmethod
    def get_poisson_probs(l_home: float, l_away: float) -> Dict:
        """
        Рассчитывает вероятности по Пуассону для различных исходов.
        """
        probs = []
        btts_yes = 0.0
        btts_no = 0.0
        over25_yes = 0.0
        over25_no = 0.0

        try:
            l_home = max(float(l_home), AnalysisConstants.POISSON_MIN_LAMBDA)
            l_away = max(float(l_away), AnalysisConstants.POISSON_MIN_LAMBDA)

            exp_home = math.exp(-l_home)
            exp_away = math.exp(-l_away)

            max_goals = AnalysisConstants.POISSON_MAX_GOALS
            factorials = [math.factorial(i) for i in range(max_goals + 1)]
            home_powers = [l_home ** i for i in range(max_goals + 1)]
            away_powers = [l_away ** i for i in range(max_goals + 1)]

            for h in range(max_goals + 1):
                p_h = (exp_home * home_powers[h]) / factorials[h]
                for a in range(max_goals + 1):
                    p_a = (exp_away * away_powers[a]) / factorials[a]
                    probability = p_h * p_a * 100

                    if probability > AnalysisConstants.MIN_PROBABILITY:
                        probs.append({
                            'score': f"{h}:{a}",
                            'prob': round(probability, 2)
                        })

                    if h > 0 and a > 0:
                        btts_yes += probability
                    else:
                        btts_no += probability

                    if (h + a) > 2.5:
                        over25_yes += probability
                    else:
                        over25_no += probability

            top_scores = sorted(probs, key=lambda x: x['prob'], reverse=True)[:5]

            total_btss = btts_yes + btts_no
            total_over = over25_yes + over25_no

            if total_btss > 0:
                btts_yes = (btts_yes / total_btss) * 100
                btts_no = (btts_no / total_btss) * 100

            if total_over > 0:
                over25_yes = (over25_yes / total_over) * 100
                over25_no = (over25_no / total_over) * 100

            btts_yes = round(btts_yes, 2)
            btts_no = round(btts_no, 2)
            over25_yes = round(over25_yes, 2)
            over25_no = round(over25_no, 2)

        except Exception:
            return {
                'top_scores': [],
                'btts_yes': 0.0,
                'btts_no': 0.0,
                'over25_yes': 0.0,
                'over25_no': 0.0
            }

        return {
            'top_scores': top_scores,
            'btts_yes': btts_yes,
            'btts_no': btts_no,
            'over25_yes': over25_yes,
            'over25_no': over25_no
        }


class CleanedTemplateView(TemplateView):
    """
    Представление для анализа матчей из Excel файла с использованием калибровочных данных.

    Что делает:
    1. Загружает матчи из for_analyze_matches.xlsx (Время, Хозяева, Гости, П1, ТБ2,5, ТМ2,5)
    2. Для каждого матча:
       - Находит команды в БД (сначала по каноническому имени, потом по алиасам)
       - Определяет лигу по последнему матчу команды
       - Рассчитывает Пуассон (лямбды и вероятность ТБ2.5) через методы модели
       - Определяет блоки: для П1, для ТБ, для вероятности (5% интервалы)
       - Ищет в калибровочных данных точное совпадение всех трех блоков
       - Получает статистику (total, hits_over) и рассчитывает hits_under = total - hits_over
       - Рассчитывает EV для ТБ и ТМ
       - Если EV > 7%, добавляет матч в результаты
    3. Сортирует результаты по времени
    4. Передает в шаблон для отображения
    """
    template_name = 'app_bets/cleaned.html'

    # Константы для блоков (точно как в скрипте анализа)
    PROBABILITY_BINS = [
        (0, 9), (10, 19), (20, 29), (30, 39), (40, 49),
        (50, 59), (60, 69), (70, 79), (80, 89), (90, 100)
    ]

    ODDS_BINS = [
        (1.00, 1.10), (1.10, 1.21), (1.21, 1.33), (1.33, 1.46), (1.46, 1.61),
        (1.61, 1.77), (1.77, 1.95), (1.95, 2.14), (2.14, 2.35), (2.35, 2.59),
        (2.59, 2.85), (2.85, 3.13), (3.13, 3.44), (3.44, 3.78), (3.78, 4.16),
        (4.16, 4.58), (4.58, 5.04), (5.04, 5.54), (5.54, 6.09), (6.09, 6.70),
        (6.70, 7.37), (7.37, 8.11), (8.11, 8.92), (8.92, 9.81), (9.81, 10.79),
        (10.79, 11.87), (11.87, 13.06), (13.06, float('inf'))
    ]

    def get_probability_bin(self, prob):
        """
        Определяет блок вероятности (5% интервалы)

        Аргументы:
            prob: вероятность в процентах (0-100)

        Возвращает:
            str: блок в формате "45-50%" или "95-100%"
        """
        for low, high in self.PROBABILITY_BINS:
            if low <= prob < high:
                return f"{low}-{high}%"
        return "95-100%"

    def get_odds_bin(self, odds):
        """
        Определяет блок коэффициента по фиксированной сетке

        Аргументы:
            odds: коэффициент (например 2.15)

        Возвращает:
            str: блок в формате "2.14-2.35" или ">13.06"
        """
        if odds is None:
            return None
        for low, high in self.ODDS_BINS:
            if low <= odds < high:
                if high == float('inf'):
                    return f">{low:.2f}"
                return f"{low:.2f}-{high:.2f}"
        return f">{self.ODDS_BINS[-1][0]:.2f}"

    def get_calibration_data(self):
        """
        Загружает калибровочные данные из PKL файла

        Ожидаемая структура:
        {
            'АПЛ Англия': {
                ('1.77-1.95', '1.95-2.14', '50-55%'): {'total': 150, 'hits': 82},
                ...
            },
            ...
        }

        Возвращает:
            dict: калибровочные данные или None при ошибке
        """
        pkl_path = os.path.join(settings.BASE_DIR, 'analysis_results', 'all_leagues_complete_stats.pkl')

        if not os.path.exists(pkl_path):
            return None

        try:
            with open(pkl_path, 'rb') as f:
                data = pickle.load(f)
            return data
        except Exception as e:
            return None

    def get_matches_from_excel(self):
        """
        Загружает матчи для анализа из Excel файла

        Ожидаемые колонки:
        - Время: время матча (22:00)
        - Хозяева: название команды хозяев
        - Гости: название команды гостей
        - П1: коэффициент на победу хозяев
        - ТБ2,5: коэффициент на тотал больше 2.5
        - ТМ2,5: коэффициент на тотал меньше 2.5

        Возвращает:
            DataFrame: данные матчей или None при ошибке
        """
        excel_path = os.path.join(settings.BASE_DIR, 'for_analyze_matches.xlsx')

        if not os.path.exists(excel_path):
            return None

        try:
            df = pd.read_excel(excel_path)
            required = ['Время', 'Хозяева', 'Гости', 'П1', 'ТБ2,5', 'ТМ2,5']

            if not all(col in df.columns for col in required):
                return None

            return df
        except Exception as e:
            return None

    def find_team(self, name):
        """
        Находит команду в БД по названию.

        Порядок поиска:
        1. Сначала по точному совпадению канонического имени (name)
        2. Затем по алиасам (если точного имени нет)
        3. Затем по частичному совпадению (как запасной вариант)

        Аргументы:
            name: название команды из Excel (например "Миллуол")

        Возвращает:
            Team: объект команды или None если не найдена
        """
        if not name:
            return None

        try:
            # Очищаем имя от лишних пробелов
            clean_name = " ".join(name.split()).lower()

            # 1. Поиск по точному каноническому имени (приоритет)
            team = Team.objects.filter(name=name).first()
            if team:
                return team

            # 2. Поиск по алиасам (если точного имени нет)
            alias = TeamAlias.objects.filter(name=clean_name).select_related('team').first()
            if alias:
                return alias.team

            # 3. Поиск по частичному совпадению (запасной вариант)
            team = Team.objects.filter(name__icontains=name).first()
            if team:
                # Создаем алиас для будущего использования
                TeamAlias.objects.get_or_create(name=clean_name, team=team)
                return team
            return None

        except Exception as e:
            return None

    def get_league_for_team(self, team):
        """
        Определяет лигу команды по ее последнему матчу в БД

        Аргументы:
            team: объект Team

        Возвращает:
            League: объект лиги или None если нет матчей
        """
        if not team:
            return None

        last_match = Match.objects.filter(
            Q(home_team=team) | Q(away_team=team)
        ).select_related('league').order_by('-date').first()

        if last_match:
            return last_match.league
        else:
            return None

    def calculate_poisson_for_match(self, home_team, away_team, league):
        """
        Рассчитывает Пуассон для матча, используя методы модели (как в AnalyzeView)
        """
        try:
            from app_bets.models import Season
            from django.utils.timezone import now

            # Получаем текущий сезон (как в AnalyzeView)
            current_season = Season.objects.filter(is_current=True).first()
            if not current_season:
                current_season = Season.objects.order_by('-start_date').first()

            # Создаем временный объект Match с правильным сезоном
            temp_match = Match(
                home_team=home_team,
                away_team=away_team,
                league=league,
                season=current_season,  # ВАЖНО: указываем сезон!
                date=now()
            )

            # Вызываем метод модели
            lambda_result = temp_match.calculate_poisson_lambda(date=now(), last_n=7)

            if 'error' in lambda_result:
                # В AnalyzeView при ошибке используются дефолтные значения
                lambda_home = 1.2
                lambda_away = 1.0
            else:
                lambda_home = lambda_result['home_lambda']
                lambda_away = lambda_result['away_lambda']

            # Получаем вероятности
            probs = AnalyzeView.get_poisson_probs(lambda_home, lambda_away)
            over_prob = probs['over25_yes']

            return {
                'home_lambda': lambda_home,
                'away_lambda': lambda_away,
                'over_prob': over_prob,
                'under_prob': 100 - over_prob
            }

        except Exception as e:
            import traceback
            traceback.print_exc()
            return None

    def find_calibration(self, calib_data, league_name, odds_h, odds_over, prob_value):
        """
        Находит калибровочные данные для матча по ПОЛНОМУ СОВПАДЕНИЮ всех трех блоков

        Возвращает данные для ТБ и автоматически рассчитывает данные для ТМ

        Аргументы:
            calib_data: словарь с калибровочными данными
            league_name: название лиги
            odds_h: коэффициент П1
            odds_over: коэффициент ТБ2.5
            prob_value: вероятность по Пуассону (в процентах)

        Возвращает:
            tuple: (over_data, under_data) где каждый элемент - словарь с ключами:
                   'prob', 'total', 'hits', 'interval' или None если нет данных
        """
        if league_name not in calib_data:
            return None, None

        league_stats = calib_data[league_name]

        # Определяем блоки точно как в скрипте анализа
        p1_bin = self.get_odds_bin(odds_h)
        tb_bin = self.get_odds_bin(odds_over)
        prob_bin = self.get_probability_bin(prob_value)

        key = (p1_bin, tb_bin, prob_bin)

        if key in league_stats:
            stats = league_stats[key]
            total = stats['total']
            hits_over = stats['hits']
            hits_under = total - hits_over

            over_prob = (hits_over / total) * 100 if total > 0 else 0
            under_prob = (hits_under / total) * 100 if total > 0 else 0

            over_data = {
                'prob': over_prob,
                'total': total,
                'hits': hits_over,
                'interval': prob_bin
            }

            under_data = {
                'prob': under_prob,
                'total': total,
                'hits': hits_under,
                'interval': prob_bin
            }

            return over_data, under_data
        else:
            return None, None

    def get_context_data(self, **kwargs):
        """
        Основной метод обработки и подготовки данных для шаблона
        """
        context = super().get_context_data(**kwargs)

        # Загружаем данные
        calib_data = self.get_calibration_data()
        excel_df = self.get_matches_from_excel()

        if calib_data is None:
            context['error'] = 'Не удалось загрузить калибровочные данные'
            return context

        if excel_df is None:
            context['error'] = 'Не удалось загрузить Excel файл с матчами'
            return context

        # Получаем параметр сортировки
        sort_param = self.request.GET.get('sort', 'time_asc')
        context['current_sort'] = sort_param

        # ========== ДИНАМИЧЕСКАЯ ГРУППИРОВКА ЛИГ ПО РЕЗУЛЬТАТИВНОСТИ ==========
        # print("\n" + "=" * 80)
        # print("ДИНАМИЧЕСКАЯ ГРУППИРОВКА ЛИГ")
        # print("=" * 80)

        # Отсортированный список лиг по результативности (из ваших данных)
        LEAGUES_WITH_SCORING = [
            ('Бундеслига Германия', 3.18),
            ('Эредивизи Нидерланды', 3.12),
            ('Бундеслига 2 Германия', 3.0),
            ('АПЛ Англия', 2.97),
            ('Суперлига Турция', 2.87),
            ('Премьер Лига Шотландия', 2.85),
            ('Лига 1 Франция', 2.83),
            ('Высшая лига Бельгия', 2.82),
            ('Высшая лига Португалия', 2.67),
            ('Ла Лига Испания', 2.6),
            ('Серия А Италия', 2.55),
            ('Чемпионшип Англия', 2.53),
            ('Лига 2 Франция', 2.48),
            ('Серия Б Италия', 2.45),
            ('Сегунда Испания', 2.31),
        ]

        # Создаем словарь для быстрого доступа к позиции лиги
        league_positions = {}
        for idx, (league_name, _) in enumerate(LEAGUES_WITH_SCORING):
            league_positions[league_name] = idx

        # print(f"Всего лиг в рейтинге: {len(LEAGUES_WITH_SCORING)}")
        # print("\nРейтинг лиг по результативности:")
        # for i, (league, avg) in enumerate(LEAGUES_WITH_SCORING):
        #     print(f"  {i + 1}. {league}: {avg:.2f}")

        # Функция для получения группы лиг для заданной лиги
        def get_league_group(league_name, neighbors=2):
            """Возвращает список лиг для группы: сама лига + neighbors выше и ниже"""
            if league_name not in league_positions:
                return [league_name]  # Если лига не найдена, возвращаем только её

            pos = league_positions[league_name]
            total = len(LEAGUES_WITH_SCORING)

            # Вычисляем диапазон с учетом границ
            start = max(0, pos - neighbors)
            end = min(total, pos + neighbors + 1)  # +1 потому что range не включает последний

            # Если не хватает сверху, добавляем снизу
            if pos - neighbors < 0:
                shortage = abs(pos - neighbors)
                end = min(total, end + shortage)

            # Если не хватает снизу, добавляем сверху
            if pos + neighbors + 1 > total:
                shortage = (pos + neighbors + 1) - total
                start = max(0, start - shortage)

            # Получаем названия лиг
            group = [LEAGUES_WITH_SCORING[i][0] for i in range(start, end)]

            return group

        # Демонстрация работы группировки
        # print("\n" + "=" * 80)
        # print("ДЕМОНСТРАЦИЯ ГРУППИРОВКИ (примеры)")
        # print("=" * 80)

        example_leagues = ['Бундеслига Германия', 'Ла Лига Испания', 'Сегунда Испания']
        for league in example_leagues:
            group = get_league_group(league, neighbors=2)
            pos = league_positions.get(league, -1) + 1
            # print(f"\nЛига: {league} (позиция {pos})")
            # print(f"  Группа ({len(group)} лиг):")
            # for g in group:
            #     print(f"    - {g}")
        # ========== КОНЕЦ ДИНАМИЧЕСКОЙ ГРУППИРОВКИ ==========

        # Анализируем матчи
        analysis_results = []
        MIN_EV = 7
        MIN_TOTAL = 5
        ALPHA = 1
        BETA = 1
        BAYES_THRESHOLD = 100

        # Статистика для отладки
        debug_stats = {
            'total_matches': 0,
            'skipped_no_odds': 0,
            'skipped_no_teams': 0,
            'skipped_no_league': 0,
            'skipped_no_poisson': 0,
            'grouping_stats': {}
        }

        # print("\n" + "=" * 80)
        # print("НАЧАЛО АНАЛИЗА МАТЧЕЙ")
        # print("=" * 80)

        for idx, row in excel_df.iterrows():
            debug_stats['total_matches'] += 1
            print(f"\n--- МАТЧ #{idx + 1} ---")

            try:
                # Парсим время
                match_time = row['Время']
                if hasattr(match_time, 'strftime'):
                    time_str = match_time.strftime('%H:%M')
                else:
                    time_str = str(match_time)

                home_name = str(row['Хозяева']).strip()
                away_name = str(row['Гости']).strip()
                odds_h = float(row['П1']) if not pd.isna(row['П1']) else None
                odds_over = float(row['ТБ2,5']) if not pd.isna(row['ТБ2,5']) else None
                odds_under = float(row['ТМ2,5']) if not pd.isna(row['ТМ2,5']) else None

                # print(
                #     f"  Данные: {time_str} | {home_name} - {away_name} | П1={odds_h}, ТБ={odds_over}, ТМ={odds_under}")

                if not odds_h or not odds_over or not odds_under:
                    debug_stats['skipped_no_odds'] += 1
                    print("  ❌ Пропущен: отсутствуют коэффициенты")
                    continue

                # Находим команды
                home_team = self.find_team(home_name)
                away_team = self.find_team(away_name)

                if not home_team:
                    debug_stats['skipped_no_teams'] += 1
                    print(f"  ❌ Пропущен: не найдена команда {home_name}")
                    continue
                if not away_team:
                    debug_stats['skipped_no_teams'] += 1
                    print(f"  ❌ Пропущен: не найдена команда {away_name}")
                    continue

                # Определяем лигу
                league = self.get_league_for_team(home_team) or self.get_league_for_team(away_team)
                if not league:
                    debug_stats['skipped_no_league'] += 1
                    print("  ❌ Пропущен: не определена лига")
                    continue

                # print(f"  ✅ Лига: {league.name}")

                # Рассчитываем Пуассон
                poisson_result = self.calculate_poisson_for_match(home_team, away_team, league)
                if not poisson_result:
                    debug_stats['skipped_no_poisson'] += 1
                    print("  ❌ Пропущен: ошибка расчета Пуассона")
                    continue

                over_prob = poisson_result['over_prob']
                under_prob = poisson_result['under_prob']
                # print(f"  📊 Пуассон: ТБ={over_prob:.1f}%, ТМ={under_prob:.1f}%")

                # Находим калибровку для КОНКРЕТНОЙ лиги
                over_data_single, under_data_single = self.find_calibration(
                    calib_data, league.name, odds_h, odds_over, over_prob
                )

                # ===== ДИНАМИЧЕСКАЯ ГРУППИРОВКА =====
                if league.name in league_positions:
                    group_leagues = get_league_group(league.name, neighbors=2)
                    league_group = f"group_{league_positions[league.name]}"  # для статистики

                    # Собираем данные из всех лиг группы
                    over_total = 0
                    over_hits = 0
                    under_total = 0
                    under_hits = 0
                    used_leagues = []

                    # print(f"  🔍 Группировка: найдено {len(group_leagues)} лиг")

                    for league_name in group_leagues:
                        over_tmp, under_tmp = self.find_calibration(
                            calib_data, league_name, odds_h, odds_over, over_prob
                        )

                        if over_tmp:
                            over_total += over_tmp['total']
                            over_hits += over_tmp['hits']
                            used_leagues.append(league_name)
                            # print(f"    + {league_name}: ТБ total={over_tmp['total']}, hits={over_tmp['hits']}")

                        if under_tmp:
                            under_total += under_tmp['total']
                            under_hits += under_tmp['hits']

                    # Формируем объединенные данные
                    over_data = None
                    under_data = None

                    if over_total > 0:
                        over_data = {
                            'total': over_total,
                            'hits': over_hits,
                            'prob': (over_hits / over_total) * 100 if over_total > 0 else 0,
                            'interval': over_data_single['interval'] if over_data_single else 'unknown'
                        }
                        # print(f"  📊 ИТОГО ТБ: total={over_total}, hits={over_hits}, prob={over_data['prob']:.1f}%")

                    if under_total > 0:
                        under_data = {
                            'total': under_total,
                            'hits': under_hits,
                            'prob': (under_hits / under_total) * 100 if under_total > 0 else 0,
                            'interval': under_data_single['interval'] if under_data_single else 'unknown'
                        }
                        # print(f"  📊 ИТОГО ТМ: total={under_total}, hits={under_hits}, prob={under_data['prob']:.1f}%")

                    # Статистика группировки
                    group_key = f"{len(group_leagues)}_лиг"
                    debug_stats['grouping_stats'][group_key] = debug_stats['grouping_stats'].get(group_key, 0) + 1
                else:
                    # Если лига не в рейтинге, используем только свои данные
                    over_data = over_data_single
                    under_data = under_data_single
                    used_leagues = [league.name] if over_data_single or under_data_single else []
                    # print(f"  ⚠️ Лига не в рейтинге, используются только свои данные")
                # ===== КОНЕЦ ГРУППИРОВКИ =====

                # Проверяем ТБ (остальная логика без изменений)
                if over_data and over_data['total'] >= MIN_TOTAL:
                    if over_data['total'] < BAYES_THRESHOLD:
                        smoothed_prob_over = (over_data['hits'] + ALPHA) / (over_data['total'] + ALPHA + BETA) * 100
                        actual_prob_used = round(smoothed_prob_over, 1)
                        smoothing_applied = True
                    else:
                        actual_prob_used = round(over_data['prob'], 1)
                        smoothing_applied = False

                    ev_over = (actual_prob_used / 100.0) * odds_over - 1
                    ev_over_percent = ev_over * 100

                    if ev_over_percent > MIN_EV:
                        result_item = {
                            'time': time_str,
                            'time_sort': time_str,
                            'home': home_name,
                            'away': away_name,
                            'match': f"{home_name} - {away_name}",
                            'league': league.name,
                            'league_group': f"group_{league_positions.get(league.name, 'unknown')}",
                            'used_leagues': used_leagues,
                            'league_sort': league.name,
                            'odds_h': odds_h,
                            'odds_over': odds_over,
                            'odds_under': odds_under,
                            'target': 'ТБ 2.5',
                            'ev': round(ev_over_percent, 1),
                            'ev_sort': ev_over_percent,
                            'poisson_prob': round(over_prob, 1),
                            'actual_prob': actual_prob_used,
                            'raw_prob': round(over_data['prob'], 1),
                            'interval': over_data['interval'],
                            'recommended_odds': odds_over,
                            'home_team_id': home_team.id,
                            'away_team_id': away_team.id,
                            'league_id': league.id,
                            'target_code': 'over',
                            'total': over_data['total'],
                            'hits': over_data['hits'],
                            'home_lambda': poisson_result['home_lambda'],
                            'away_lambda': poisson_result['away_lambda'],
                            'smoothing_applied': smoothing_applied,
                        }
                        analysis_results.append(result_item)
                        # print(f"  ✅ ТБ ДОБАВЛЕН! EV={ev_over_percent:.1f}%")

                # Проверяем ТМ (аналогично)
                if under_data and under_data['total'] >= MIN_TOTAL:
                    if under_data['total'] < BAYES_THRESHOLD:
                        smoothed_prob_under = (under_data['hits'] + ALPHA) / (under_data['total'] + ALPHA + BETA) * 100
                        actual_prob_used = round(smoothed_prob_under, 1)
                        smoothing_applied = True
                    else:
                        actual_prob_used = round(under_data['prob'], 1)
                        smoothing_applied = False

                    ev_under = (actual_prob_used / 100.0) * odds_under - 1
                    ev_under_percent = ev_under * 100

                    if ev_under_percent > MIN_EV:
                        result_item = {
                            'time': time_str,
                            'time_sort': time_str,
                            'home': home_name,
                            'away': away_name,
                            'match': f"{home_name} - {away_name}",
                            'league': league.name,
                            'league_group': f"group_{league_positions.get(league.name, 'unknown')}",
                            'used_leagues': used_leagues,
                            'league_sort': league.name,
                            'odds_h': odds_h,
                            'odds_over': odds_over,
                            'odds_under': odds_under,
                            'target': 'ТМ 2.5',
                            'ev': round(ev_under_percent, 1),
                            'ev_sort': ev_under_percent,
                            'poisson_prob': round(under_prob, 1),
                            'actual_prob': actual_prob_used,
                            'raw_prob': round(under_data['prob'], 1),
                            'interval': under_data['interval'],
                            'recommended_odds': odds_under,
                            'home_team_id': home_team.id,
                            'away_team_id': away_team.id,
                            'league_id': league.id,
                            'target_code': 'under',
                            'total': under_data['total'],
                            'hits': under_data['hits'],
                            'home_lambda': poisson_result['home_lambda'],
                            'away_lambda': poisson_result['away_lambda'],
                            'smoothing_applied': smoothing_applied,
                        }
                        analysis_results.append(result_item)
                        # print(f"  ✅ ТМ ДОБАВЛЕН! EV={ev_under_percent:.1f}%")

            except Exception as e:
                import traceback
                print(f"❌ Ошибка при обработке матча #{idx + 1}: {e}")
                traceback.print_exc()
                continue

        # ИТОГОВАЯ СТАТИСТИКА
        # print("\n" + "=" * 80)
        # print("ИТОГОВАЯ СТАТИСТИКА")
        # print("=" * 80)
        #
        # print(f"\n📊 Всего обработано матчей: {debug_stats['total_matches']}")
        # print(f"📊 Найдено сигналов: {len(analysis_results)}")
        #
        # print("\n🔍 Пропущено матчей:")
        # print(f"  ❌ Нет коэффициентов: {debug_stats['skipped_no_odds']}")
        # print(f"  ❌ Нет команд: {debug_stats['skipped_no_teams']}")
        # print(f"  ❌ Нет лиги: {debug_stats['skipped_no_league']}")
        # print(f"  ❌ Нет Пуассона: {debug_stats['skipped_no_poisson']}")
        #
        # print("\n📊 Статистика группировки:")
        # for group_key, count in debug_stats['grouping_stats'].items():
        #     print(f"  {group_key}: {count} матчей")

        # Применяем сортировку
        if sort_param == 'time_asc':
            analysis_results.sort(key=lambda x: x['time_sort'])
        elif sort_param == 'time_desc':
            analysis_results.sort(key=lambda x: x['time_sort'], reverse=True)
        elif sort_param == 'league_asc':
            analysis_results.sort(key=lambda x: x['league_sort'])
        elif sort_param == 'league_desc':
            analysis_results.sort(key=lambda x: x['league_sort'], reverse=True)
        elif sort_param == 'ev_asc':
            analysis_results.sort(key=lambda x: x['ev_sort'])
        elif sort_param == 'ev_desc':
            analysis_results.sort(key=lambda x: x['ev_sort'], reverse=True)

        # Сохраняем в сессию
        self.request.session['cleaned_analysis_results'] = analysis_results

        context['analysis_results'] = analysis_results
        context['total_analyzed'] = len(analysis_results)
        context['min_ev'] = MIN_EV

        return context


class UploadCSVView(View):
    template_name = 'app_bets/bets_main.html'

    def post(self, request):
        context = {
            'results': [],
            'raw_text': '',
            'unknown_teams': [],
            'all_teams': Team.objects.all(),
            'import_status': 'success',
            'import_message': '',
            'import_added': 0,
            'import_skipped': 0,
            'import_errors': 0
        }

        try:
            if 'csv_file' not in request.FILES:
                if 'sync_files' in request.POST:
                    return self.sync_from_folder(request, context)

                context['import_status'] = 'error'
                context['import_message'] = 'Файл не найден. Выберите CSV файл для загрузки.'
                return render(request, self.template_name, context)

            csv_file = request.FILES['csv_file']

            if not csv_file.name.endswith('.csv'):
                context['import_status'] = 'error'
                context['import_message'] = 'Файл должен быть в формате CSV'
                return render(request, self.template_name, context)

            return self.import_from_file(request, csv_file, context)

        except Exception as e:
            context['import_status'] = 'error'
            context['import_message'] = f'Ошибка при обработке запроса: {str(e)}'
            return render(request, self.template_name, context)

    def sync_from_folder(self, request, context):
        import_data_dir = os.path.join(settings.BASE_DIR, 'import_data')

        try:
            if not os.path.exists(import_data_dir):
                context['import_status'] = 'error'
                context['import_message'] = f'Папка {import_data_dir} не найдена.'
                return render(request, self.template_name, context)

            csv_files = [f for f in os.listdir(import_data_dir) if f.endswith('.csv')]

            if not csv_files:
                context['import_status'] = 'warning'
                context['import_message'] = f'В папке {import_data_dir} не найдено CSV файлов.'
                return render(request, self.template_name, context)

            total_added = 0
            total_skipped = 0
            total_errors = 0
            total_aliases = 0
            processed_files = 0
            details = []
            all_unknown_teams = set()

            self.request = request

            for csv_file_name in csv_files:
                file_path = os.path.join(import_data_dir, csv_file_name)
                result = self.process_csv_file(file_path)

                total_added += result['added']
                total_skipped += result['skipped']
                total_errors += result['errors']
                total_aliases += result.get('created_aliases', 0)
                processed_files += 1

                if 'unknown_teams_list' in result:
                    for team in result['unknown_teams_list']:
                        all_unknown_teams.add(team['name'])

                details.append(
                    f"{csv_file_name}: +{result['added']} "
                    f"(пропущено {result['skipped']}, "
                    f"ошибок {result['errors']}, "
                    f"алиасов {result.get('created_aliases', 0)})"
                )

            if all_unknown_teams:
                current_unknown = request.session.get('unknown_teams', [])
                request.session['unknown_teams'] = list(set(current_unknown + list(all_unknown_teams)))

            context['import_added'] = total_added
            context['import_skipped'] = total_skipped
            context['import_errors'] = total_errors
            context['import_aliases'] = total_aliases

            message_parts = [
                f'СИНХРОНИЗАЦИЯ ЗАВЕРШЕНА:',
                f'✅ Обработано файлов: {processed_files}',
                f'✅ Добавлено матчей: {total_added}',
                f'⚠️ Пропущено: {total_skipped}',
                f'❌ Ошибок: {total_errors}',
                f'✨ Создано алиасов: {total_aliases}',
            ]

            if all_unknown_teams:
                message_parts.append(f'\n📝 Неизвестных команд: {len(all_unknown_teams)}')

            message_parts.append(f'\n📊 Детали по файлам:')
            message_parts.extend(details)

            context['import_message'] = '\n'.join(message_parts)

            if total_added == 0 and total_aliases == 0:
                context['import_status'] = 'warning'
            elif total_added > 0:
                context['import_status'] = 'success'
            else:
                context['import_status'] = 'info'

        except Exception as e:
            context['import_status'] = 'error'
            context['import_message'] = f'Ошибка синхронизации: {str(e)}'

        return render(request, self.template_name, context)

    def import_from_file(self, request, csv_file, context):
        try:
            try:
                file_content = csv_file.read().decode('utf-8-sig')
            except UnicodeDecodeError:
                csv_file.seek(0)
                file_content = csv_file.read().decode('latin-1')

            import tempfile

            with tempfile.NamedTemporaryFile(mode='w', encoding='utf-8', suffix='.csv', delete=False) as tmp:
                tmp.write(file_content)
                tmp_path = tmp.name

            try:
                result = self.process_csv_file(tmp_path)
            finally:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)

            context['import_added'] = result['added']
            context['import_skipped'] = result['skipped']
            context['import_errors'] = result['errors']

            if result['added'] > 0:
                context['import_message'] = (
                    f'ИМПОРТ ЗАВЕРШЕН:\n'
                    f'- Добавлено матчей: {result["added"]}\n'
                    f'- Пропущено: {result["skipped"]}\n'
                    f'- Ошибок: {result["errors"]}'
                )
            else:
                context['import_status'] = 'warning'
                context['import_message'] = (
                    f'Нет новых матчей для добавления.\n'
                    f'Пропущено: {result["skipped"]}\n'
                    f'Ошибок: {result["errors"]}'
                )

        except Exception as e:
            context['import_status'] = 'error'
            context['import_message'] = f'Ошибка при обработке файла: {str(e)}'

        return render(request, self.template_name, context)

    @transaction.atomic
    def process_csv_file(self, file_path):
        count = 0
        skipped = 0
        errors = 0
        processed = 0
        created_aliases = 0
        unknown_teams_list = []

        all_teams = {team.id: team for team in Team.objects.all()}
        all_teams_by_name = {team.name.lower(): team for team in Team.objects.all()}
        all_aliases = {}
        for alias in TeamAlias.objects.all().select_related('team'):
            all_aliases[alias.name] = alias.team
        all_leagues = {league.name: league for league in League.objects.all()}

        def find_team_smart(team_name, all_teams_dict, all_aliases_dict, all_teams_by_name_dict):
            nonlocal created_aliases, unknown_teams_list

            if not team_name:
                return None

            clean_name = self.clean_team_name(team_name)

            if clean_name in all_aliases_dict:
                return all_aliases_dict[clean_name]

            if clean_name in all_teams_by_name_dict:
                team = all_teams_by_name_dict[clean_name]
                alias, created = TeamAlias.objects.get_or_create(
                    name=clean_name,
                    defaults={'team': team}
                )
                if created:
                    created_aliases += 1
                    all_aliases_dict[clean_name] = team
                return team

            best_match = None
            best_score = 0

            for team_name_db, team in all_teams_by_name_dict.items():
                if clean_name in team_name_db:
                    score = len(clean_name) / len(team_name_db)
                    if score > best_score:
                        best_score = score
                        best_match = team
                elif team_name_db in clean_name:
                    score = len(team_name_db) / len(clean_name)
                    if score > best_score:
                        best_score = score
                        best_match = team

            if best_match and best_score > 0.6:
                alias, created = TeamAlias.objects.get_or_create(
                    name=clean_name,
                    defaults={'team': best_match}
                )
                if created:
                    created_aliases += 1
                    all_aliases_dict[clean_name] = best_match
                return best_match

            unknown_teams_list.append({
                'name': team_name,
                'clean_name': clean_name
            })
            return None

        try:
            with open(file_path, mode='r', encoding='utf-8-sig') as f:
                first_line = f.readline()
                f.seek(0)

                delimiter = ';' if ';' in first_line else ','
                reader = csv.DictReader(f, delimiter=delimiter)

                for row in reader:
                    processed += 1
                    try:
                        div_code = row.get('Div', '').strip()
                        if not div_code:
                            skipped += 1
                            continue

                        league_name = ParsingConstants.DIV_TO_LEAGUE_NAME.get(div_code)
                        if not league_name:
                            skipped += 1
                            continue

                        league = all_leagues.get(league_name)
                        if not league:
                            skipped += 1
                            continue

                        date_str = row.get('Date', '').strip()
                        if not date_str:
                            skipped += 1
                            continue

                        try:
                            dt = datetime.strptime(date_str, '%d/%m/%Y')
                        except ValueError:
                            try:
                                dt = datetime.strptime(date_str, '%d/%m/%y')
                            except ValueError:
                                try:
                                    dt = datetime.strptime(date_str, '%Y-%m-%d')
                                except ValueError:
                                    errors += 1
                                    continue

                        season = self.get_season_by_date(dt)
                        if not season:
                            season = Season.objects.filter(
                                start_date__lte=dt.date(),
                                end_date__gte=dt.date()
                            ).first()
                            if not season:
                                errors += 1
                                continue

                        home_team_name = row.get('HomeTeam', '').strip()
                        away_team_name = row.get('AwayTeam', '').strip()

                        if not home_team_name or not away_team_name:
                            skipped += 1
                            continue

                        home_team = find_team_smart(home_team_name, all_teams, all_aliases, all_teams_by_name)
                        away_team = find_team_smart(away_team_name, all_teams, all_aliases, all_teams_by_name)

                        if not home_team or not away_team:
                            skipped += 1
                            continue

                        dt_aware = make_aware(dt, get_current_timezone())

                        if Match.objects.filter(
                                date=dt_aware,
                                home_team=home_team,
                                away_team=away_team
                        ).exists():
                            skipped += 1
                            continue

                        odd_h = self.parse_odd(row.get('AvgH') or row.get('B365H') or row.get('PSH') or '1.01')
                        odd_d = self.parse_odd(row.get('AvgD') or row.get('B365D') or row.get('PSD') or '1.01')
                        odd_a = self.parse_odd(row.get('AvgA') or row.get('B365A') or row.get('PSA') or '1.01')
                        h_goal = self.parse_score(row.get('FTHG') or '0')
                        a_goal = self.parse_score(row.get('FTAG') or '0')

                        Match.objects.create(
                            season=season,
                            league=league,
                            date=dt_aware,
                            home_team=home_team,
                            away_team=away_team,
                            home_score_reg=h_goal,
                            away_score_reg=a_goal,
                            home_score_final=h_goal,
                            away_score_final=a_goal,
                            odds_home=odd_h,
                            odds_draw=odd_d,
                            odds_away=odd_a,
                            finish_type='REG'
                        )

                        count += 1

                    except Exception:
                        errors += 1
                        continue

        except Exception:
            errors += 1

        if unknown_teams_list and hasattr(self, 'request'):
            current_unknown = self.request.session.get('unknown_teams', [])
            new_unknown = list(set([item['name'] for item in unknown_teams_list]))
            self.request.session['unknown_teams'] = list(set(current_unknown + new_unknown))

        return {
            'added': count,
            'skipped': skipped,
            'errors': errors,
            'created_aliases': created_aliases,
            'unknown_teams_list': unknown_teams_list,
            'unknown_teams': len(unknown_teams_list)
        }

    @staticmethod
    def get_team_by_alias(name):
        if not name:
            return None
        clean_alias = " ".join(str(name).split()).lower()
        alias = TeamAlias.objects.filter(name=clean_alias).select_related('team').first()
        return alias.team if alias else None

    @staticmethod
    def get_season_by_date(dt):
        return Season.objects.filter(start_date__lte=dt.date(), end_date__gte=dt.date()).first()

    @staticmethod
    def parse_score(val):
        if not val or str(val).strip() == "" or str(val).lower() == 'nan':
            return 0
        try:
            return int(float(str(val).replace(',', '.')))
        except (ValueError, TypeError):
            return 0

    @staticmethod
    def parse_odd(val):
        if not val or str(val).strip() == "" or str(val).lower() == 'nan':
            return Decimal('1.01')
        try:
            return Decimal(str(val).replace(',', '.')).quantize(Decimal('0.01'))
        except:
            return Decimal('1.01')

    @staticmethod
    def clean_team_name(name: str) -> str:
        if not name:
            return ""
        clean = " ".join(str(name).split()).lower()
        clean = re.sub(r'[^\w\s]', '', clean)
        return clean


class ExportBetsExcelView(View):
    """
    Экспорт отфильтрованных и отсортированных результатов в Excel.
    """

    def get(self, request, *args, **kwargs):
        # Получаем результаты из сессии (уже отсортированные)
        results = request.session.get('results', [])
        current_sort = request.session.get('current_sort', 'default')

        # Определяем название сортировки
        sort_names = {
            'default': 'По умолчанию',
            'btts_desc': 'ОЗ (убывание)',
            'over25_desc': 'б2.5 (убывание)',
            'twins_p1_desc': 'Близнецы (макс. П1/П2)',
            'pattern_p1_desc': 'История (макс. П1/П2)'
        }
        sort_name = sort_names.get(current_sort, 'По умолчанию')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Анализ матчей"

        # Заголовок с информацией о сортировке
        ws.append([f"Сортировка: {sort_name}"])
        ws.append([])

        # Заголовки таблицы - оптимизированные, без дублей
        headers = [
            'Хозяева',
            'Гости',
            'Лига',
            'П1',
            'X',
            'П2',
            'ОЗ Да',
            'Тот2.5 Да',
            'Ист ТБ',
            'Синтез ТБ',
            'Оценка',
            'Близ П1',
            'Близ X',
            'Близ П2',
            'Ист П1',
            'Ист X',
            'Ист П2',
        ]
        ws.append(headers)

        # Жирный шрифт для заголовков
        for cell in ws[2]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Заполняем данными
        for res in results:
            # Разбиваем match на хозяева и гости
            match_parts = res['match'].split(' - ', 1)
            home_team = match_parts[0] if len(match_parts) > 0 else ''
            away_team = match_parts[1] if len(match_parts) > 1 else ''

            # Коэффициенты
            odds = res.get('odds', (None, None, None))

            # Данные исторического тотала
            historical_total = res.get('historical_total', {})
            historical_tb = f"{historical_total.get('over_25', '')}%" if historical_total and historical_total.get(
                'over_25') else ''

            # Данные Пуассона по тоталу
            poisson_over25 = res.get('poisson_over25', {})
            poisson_tb = poisson_over25.get('yes', 0)

            # ------------------------------------------------------------
            # СИНТЕТИЧЕСКИЙ ПОКАЗАТЕЛЬ
            # ------------------------------------------------------------
            synthesis_tb = ''
            confidence = ''

            if historical_total and poisson_tb:
                hist_prob = historical_total.get('over_25', 0)

                if hist_prob:
                    # 1. БАЗОВЫЙ СИГНАЛ (приоритет Пуассона 60% / История 40%)
                    base = poisson_tb * 0.6 + hist_prob * 0.4

                    # 2. БОНУС ЗА СОГЛАСОВАННОСТЬ
                    if poisson_tb > 50 and hist_prob > 50:
                        p1 = poisson_tb / 100
                        p2 = hist_prob / 100
                        boost = 1.0 + (p1 * p2 * 0.3)
                        final = base * boost
                    else:
                        final = base

                    # 3. ШТРАФ ЗА ПРОТИВОРЕЧИЕ
                    if (poisson_tb > 70 and hist_prob < 40) or (hist_prob > 70 and poisson_tb < 40):
                        gap = abs(poisson_tb - hist_prob) / 100
                        penalty = 1.0 - (gap * 0.3)
                        final = final * penalty

                    # Ограничиваем шкалу 0-100
                    final = max(0, min(100, final))

                    synthesis_tb = f"{round(final, 1)}%"

                    # 4. УРОВЕНЬ УВЕРЕННОСТИ
                    if final >= 85:
                        confidence = "ВЫСОЧАЙШАЯ"
                    elif final >= 75:
                        confidence = "ВЫСОКАЯ"
                    elif final >= 65:
                        confidence = "ВЫШЕ СРЕДНЕГО"
                    elif final >= 55:
                        confidence = "СРЕДНЯЯ"
                    elif final >= 45:
                        confidence = "НИЗКАЯ"
                    else:
                        confidence = "СЛУЧАЙНАЯ"

            # Данные близнецов
            twins = res.get('twins_data', {})
            twins_p1 = twins.get('p1', '') if twins else ''
            twins_x = twins.get('x', '') if twins else ''
            twins_p2 = twins.get('p2', '') if twins else ''

            # Данные паттернов
            pattern = res.get('pattern_data', {})
            pattern_p1 = pattern.get('p1', '') if pattern else ''
            pattern_x = pattern.get('x', '') if pattern else ''
            pattern_p2 = pattern.get('p2', '') if pattern else ''

            row = [
                home_team,
                away_team,
                res.get('league', ''),
                odds[0] if odds[0] is not None else '',
                odds[1] if odds[1] is not None else '',
                odds[2] if odds[2] is not None else '',
                res.get('poisson_btts', {}).get('yes', ''),  # ОЗ Да
                poisson_over25.get('yes', ''),  # Тотал >2.5 Да
                historical_tb,  # Ист ТБ
                synthesis_tb,  # Синтез ТБ
                confidence,  # Оценка
                f"{twins_p1}%" if twins_p1 != '' else '',  # Близнецы (П1)
                f"{twins_x}%" if twins_x != '' else '',  # Близнецы (X)
                f"{twins_p2}%" if twins_p2 != '' else '',  # Близнецы (П2)
                f"{pattern_p1}%" if pattern_p1 != '' else '',  # История (П1)
                f"{pattern_x}%" if pattern_x != '' else '',  # История (X)
                f"{pattern_p2}%" if pattern_p2 != '' else '',  # История (П2)
            ]
            ws.append(row)

        # Настройка ширины колонок
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response[
            'Content-Disposition'] = f'attachment; filename=bets_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response


class ExportCleanedExcelView(View):
    def get(self, request, *args, **kwargs):
        results = request.session.get('cleaned_analysis_results', [])
        if not results:
            # Если нет данных, можно вернуть пустой файл или ошибку
            return HttpResponse("Нет данных для экспорта", status=404)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Анализ матчей"

        # Устанавливаем шрифт Times New Roman для всей книги
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # Создаем стили
        times_new_roman = Font(name='Times New Roman', size=14)
        bold_times = Font(name='Times New Roman', size=14, bold=True)

        # Центрирование для всех ячеек
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Заголовки колонок
        headers = [
            'Время',
            'Хозяева',
            'Гости',
            'Лига',
            'Коэффициент',
            'Исход',
            'Прогноз\nПуассона,%',
            'Фактическая\nвероятность,%',
            'EV,\n%',
            'Всего\nсобытий',
            'Попаданий'
        ]
        ws.append(headers)

        # Устанавливаем ширину колонок
        column_widths = [10, 20, 20, 25, 12, 12, 18, 23, 10, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Форматируем заголовки
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name='Times New Roman', size=14, bold=True, color="FFFFFF")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Данные
        for row_num, res in enumerate(results, 2):
            row = [
                res.get('time', ''),
                res.get('home', ''),
                res.get('away', ''),
                res.get('league', ''),
                res.get('recommended_odds', ''),
                res.get('target', ''),
                res.get('poisson_prob', ''),
                res.get('actual_prob', ''),
                res.get('ev', ''),
                res.get('total', ''),
                res.get('hits', ''),
            ]
            ws.append(row)

            # Форматируем каждую ячейку строки
            for col in range(1, len(row) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font = times_new_roman
                cell.alignment = center_alignment
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Для числовых колонок устанавливаем формат
                if col in [5, 7, 8, 9]:  # Коэффициент, вероятности, EV
                    cell.number_format = '0.00'
                elif col in [10, 11]:  # Всего событий, Попаданий
                    cell.number_format = '0'  # Целые числа без десятичных знаков

        # Добавляем фильтры на заголовки
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"

        # Фиксируем заголовки (чтобы всегда были видны)
        ws.freeze_panes = 'A2'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"cleaned_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response


class TeamAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = Team.objects.all()
        if self.q:
            # Поиск только по основному имени команды (без алиасов)
            qs = qs.filter(name__istartswith=self.q)
        return qs


class LeagueAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = League.objects.all()
        if self.q:
            qs = qs.filter(name__istartswith=self.q)
        return qs


class BetCreateView(SuccessMessageMixin, CreateView):
    model = Bet
    form_class = BetForm
    template_name = 'app_bets/bet_form.html'
    success_url = reverse_lazy('app_bets:cleaned')
    success_message = "Ставка на матч %(home_team)s - %(away_team)s (коэф. %(recommended_odds)s, сумма %(stake)s) успешно сохранена!"

    def get_initial(self):
        initial = super().get_initial()

        def safe_float(val):
            if val is None:
                return None
            val = str(val).replace(',', '.')
            try:
                return float(val)
            except (TypeError, ValueError):
                return None

        def safe_int(val):
            if val is None:
                return None
            try:
                return int(val)
            except (TypeError, ValueError):
                return None

        # ID команд и лиги
        home_id = self.request.GET.get('home_team_id')
        if home_id:
            initial['home_team'] = safe_int(home_id)

        away_id = self.request.GET.get('away_team_id')
        if away_id:
            initial['away_team'] = safe_int(away_id)

        league_id = self.request.GET.get('league_id')
        if league_id:
            initial['league'] = safe_int(league_id)

        # Время матча
        match_time = self.request.GET.get('match_time')
        if match_time:
            initial['match_time'] = match_time

        # Коэффициенты (скрытые)
        odds_over = self.request.GET.get('odds_over')
        if odds_over:
            initial['odds_over'] = safe_float(odds_over)

        odds_under = self.request.GET.get('odds_under')
        if odds_under:
            initial['odds_under'] = safe_float(odds_under)

        # Рекомендуемый исход и коэффициент
        recommended_target = self.request.GET.get('recommended_target')
        if recommended_target:
            initial['recommended_target'] = recommended_target

        recommended_odds = self.request.GET.get('recommended_odds')
        if recommended_odds:
            initial['recommended_odds'] = safe_float(recommended_odds)

        # Вероятности и EV
        poisson_prob = self.request.GET.get('poisson_prob')
        if poisson_prob:
            initial['poisson_prob'] = safe_float(poisson_prob)

        actual_prob = self.request.GET.get('actual_prob')
        if actual_prob:
            initial['actual_prob'] = safe_float(actual_prob)

        ev = self.request.GET.get('ev')
        if ev:
            initial['ev'] = safe_float(ev)

        # n_last_matches (скрытое)
        n_last_matches = self.request.GET.get('n_last_matches')
        if n_last_matches:
            initial['n_last_matches'] = safe_int(n_last_matches)

        # Интервал
        interval = self.request.GET.get('interval')
        if interval:
            initial['interval'] = interval

        # Статистика выборки
        hits = self.request.GET.get('hits')
        if hits:
            initial['hits'] = safe_int(hits)

        total = self.request.GET.get('total')
        if total:
            initial['total'] = safe_int(total)

        # Автоматические поля
        from .models import Bank
        from django.utils import timezone
        initial['bank_before'] = float(Bank.get_balance())
        initial['settled_at'] = timezone.now().date().isoformat()
        initial['fractional_kelly'] = 0.25
        # ИЗМЕНЕНО С 0.5 НА 0.25

        # Расчёт начальной суммы ставки
        try:
            bank = initial.get('bank_before', 0)
            odds = initial.get('recommended_odds', 0)
            prob = initial.get('actual_prob', 0)  # это в процентах
            fraction = initial.get('fractional_kelly', 0.25)  # УЖЕ 0.25

            if odds and prob and bank and fraction and odds > 1:
                # Переводим проценты в доли
                p = float(prob) / 100

                # Критерий Келли: f = (p * k - 1) / (k - 1)
                full_kelly = (p * float(odds) - 1) / (float(odds) - 1)

                # Ограничиваем от 0 до 1
                limited_kelly = max(0, min(full_kelly, 1))

                # Применяем дробный Келли
                stake = float(bank) * limited_kelly * float(fraction)

                # Округляем до сотен
                initial['stake'] = round(stake / 100) * 100
            else:
                initial['stake'] = 0
        except Exception as e:
            print(f"Ошибка расчёта stake: {e}")
            initial['stake'] = 0

        return initial

    def get_success_message(self, cleaned_data):
        """Формирует сообщение об успешном сохранении"""
        return self.success_message % {
            'home_team': self.object.home_team.name if self.object.home_team else '?',
            'away_team': self.object.away_team.name if self.object.away_team else '?',
            'recommended_odds': self.object.recommended_odds,
            'stake': self.object.stake,
        }

    def form_invalid(self, form):
        """Обработка невалидной формы"""
        messages.error(self.request, 'Ошибка при сохранении ставки. Пожалуйста, исправьте ошибки в форме.')
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        """Добавляет сообщения в контекст"""
        context = super().get_context_data(**kwargs)

        # Добавляем сообщения из GET-параметров (если есть)
        if 'message' in self.request.GET:
            messages.success(self.request, self.request.GET.get('message'))
        if 'error' in self.request.GET:
            messages.error(self.request, self.request.GET.get('error'))

        return context


class BetRecordsView(LoginRequiredMixin, ListView):
    """
    View для отображения записей ставок в стиле админки
    """
    model = Bet
    template_name = 'app_bets/bet_records.html'
    context_object_name = 'bets'
    paginate_by = 20

    def get_queryset(self):
        """
        Получение queryset с фильтрацией и сортировкой
        """
        queryset = Bet.objects.select_related(
            'home_team',
            'away_team',
            'league',
        )

        # Фильтры
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(home_team__name__icontains=search_query) |
                Q(away_team__name__icontains=search_query) |
                Q(league__name__icontains=search_query)
            )

        date_from = self.request.GET.get('date_from', '')
        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(date_placed__date__gte=date_from)
            except ValueError:
                pass

        date_to = self.request.GET.get('date_to', '')
        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
                date_to = datetime.combine(date_to, datetime.max.time())
                queryset = queryset.filter(date_placed__lte=date_to)
            except ValueError:
                pass

        league_id = self.request.GET.get('league', '')
        if league_id and league_id.isdigit():
            queryset = queryset.filter(league_id=league_id)

        result = self.request.GET.get('result', '')
        if result:
            queryset = queryset.filter(result=result)

        # Сортировка
        sort_field = self.request.GET.get('sort', '-date_placed')
        valid_sort_fields = [
            'date_placed', '-date_placed',
            'stake', '-stake',
            'ev', '-ev',
            'profit', '-profit',
            'recommended_odds', '-recommended_odds',
        ]

        if sort_field in valid_sort_fields:
            queryset = queryset.order_by(sort_field)
        else:
            queryset = queryset.order_by('-date_placed')

        return queryset

    def get_context_data(self, **kwargs):
        """
        Добавление дополнительных данных в контекст
        """
        context = super().get_context_data(**kwargs)

        # Текущий баланс - берем из БД
        from .models import Bank
        context['current_balance'] = Bank.get_balance()

        # ВСЕ СТАВКИ для статистики
        all_bets = Bet.objects.all()

        # Сумма ставок
        total_stake_agg = all_bets.aggregate(total=Sum('stake'))['total']
        context['total_stake'] = total_stake_agg if total_stake_agg else 0

        # Прибыль
        total_profit_agg = all_bets.aggregate(total=Sum('profit'))['total']
        context['total_profit'] = total_profit_agg if total_profit_agg else 0

        # Количество ставок
        context['total_bets'] = all_bets.count()
        context['wins_count'] = all_bets.filter(result=Bet.ResultChoices.WIN).count()
        context['losses_count'] = all_bets.filter(result=Bet.ResultChoices.LOSS).count()
        context['refunds_count'] = all_bets.filter(result=Bet.ResultChoices.REFUND).count()

        # ROI
        if context['total_stake'] > 0:
            context['roi'] = (context['total_profit'] / context['total_stake']) * 100
        else:
            context['roi'] = 0

        # Список лиг для фильтра
        context['leagues'] = League.objects.filter(bet__isnull=False).distinct().order_by('name')

        return context


class StatsView(TemplateView):
    template_name = 'app_bets/stats.html'

    # Константы из скрипта анализа
    PROBABILITY_BINS = [
        (0, 9), (10, 19), (20, 29), (30, 39), (40, 49),
        (50, 59), (60, 69), (70, 79), (80, 89), (90, 100)
    ]

    ODDS_BINS = [
        (1.00, 1.10), (1.10, 1.21), (1.21, 1.33), (1.33, 1.46), (1.46, 1.61),
        (1.61, 1.77), (1.77, 1.95), (1.95, 2.14), (2.14, 2.35), (2.35, 2.59),
        (2.59, 2.85), (2.85, 3.13), (3.13, 3.44), (3.44, 3.78), (3.78, 4.16),
        (4.16, 4.58), (4.58, 5.04), (5.04, 5.54), (5.54, 6.09), (6.09, 6.70),
        (6.70, 7.37), (7.37, 8.11), (8.11, 8.92), (8.92, 9.81), (9.81, 10.79),
        (10.79, 11.87), (11.87, 13.06), (13.06, float('inf'))
    ]

    def get_probability_bins_list(self):
        """Возвращает список строк для блоков вероятности"""
        bins = []
        for low, high in self.PROBABILITY_BINS:
            bins.append(f"{low}-{high}%")
        return bins

    def get_odds_bins_list(self):
        """Возвращает список строк для блоков коэффициентов"""
        bins = []
        for low, high in self.ODDS_BINS:
            if high == float('inf'):
                bins.append(f">{low:.2f}")
            else:
                bins.append(f"{low:.2f}-{high:.2f}")
        return bins

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Загружаем данные из PKL файла
        import os
        import pickle
        from django.conf import settings

        pkl_path = os.path.join(settings.BASE_DIR, 'analysis_results', 'all_leagues_complete_stats.pkl')

        # Списки всех возможных блоков
        context['prob_bins'] = self.get_probability_bins_list()
        context['p1_bins'] = self.get_odds_bins_list()
        context['tb_bins'] = self.get_odds_bins_list()

        # Варианты сортировки для топ-10
        context['sort_options'] = [
            {'value': 'prob_bin', 'label': 'По блоку вероятности'},
            {'value': 'total', 'label': 'По количеству событий'},
            {'value': 'hits', 'label': 'По количеству попаданий'},
            {'value': 'hit_rate', 'label': 'По проценту попаданий'},
        ]

        # Варианты направления
        context['direction_options'] = [
            {'value': 'asc', 'label': 'Первые 10 (по возрастанию)'},
            {'value': 'desc', 'label': 'Последние 10 (по убыванию)'},
        ]

        # Варианты минимального количества событий
        context['min_total_options'] = [
            {'value': '1', 'label': 'Любое (≥1)'},
            {'value': '5', 'label': '≥5 событий'},
            {'value': '10', 'label': '≥10 событий'},
            {'value': '20', 'label': '≥20 событий'},
            {'value': '30', 'label': '≥30 событий'},
            {'value': '50', 'label': '≥50 событий'},
            {'value': '100', 'label': '≥100 событий'},
        ]

        if os.path.exists(pkl_path):
            try:
                with open(pkl_path, 'rb') as f:
                    data = pickle.load(f)

                # Получаем список всех лиг
                context['leagues'] = sorted(data.keys())

                # Получаем параметры из GET
                tab = self.request.GET.get('tab', 'search')
                context['current_tab'] = tab

                if tab == 'search':
                    # Режим поиска по конкретным блокам
                    selected_league = self.request.GET.get('league')
                    p1_bin = self.request.GET.get('p1_bin')
                    tb_bin = self.request.GET.get('tb_bin')
                    prob_bin = self.request.GET.get('prob_bin')

                    context['selected_league'] = selected_league
                    context['selected_p1_bin'] = p1_bin
                    context['selected_tb_bin'] = tb_bin
                    context['selected_prob_bin'] = prob_bin

                    # Если выбрана лига и все блоки, ищем данные
                    if selected_league and p1_bin and tb_bin and prob_bin:
                        if selected_league in data:
                            key = (p1_bin, tb_bin, prob_bin)
                            stats = data[selected_league].get(key)

                            if stats:
                                context['stats'] = stats
                                context['key_found'] = True
                                # Рассчитываем процент попаданий
                                if stats['total'] > 0:
                                    context['hit_rate'] = (stats['hits'] / stats['total']) * 100
                            else:
                                context['key_found'] = False

                elif tab == 'top':
                    # Режим топ-10 по лиге
                    selected_league = self.request.GET.get('league')
                    sort_by = self.request.GET.get('sort_by', 'total')
                    direction = self.request.GET.get('direction', 'desc')
                    min_total = int(self.request.GET.get('min_total', 1))
                    limit = 10

                    context['selected_league'] = selected_league
                    context['selected_sort'] = sort_by
                    context['selected_direction'] = direction
                    context['selected_min_total'] = str(min_total)

                    if selected_league and selected_league in data:
                        league_data = data[selected_league]

                        # Преобразуем данные в список для сортировки
                        results_list = []
                        for key, stats in league_data.items():
                            p1_bin, tb_bin, prob_bin = key

                            # Применяем фильтр по минимальному количеству
                            if stats['total'] < min_total:
                                continue

                            hit_rate = (stats['hits'] / stats['total'] * 100) if stats['total'] > 0 else 0

                            results_list.append({
                                'p1_bin': p1_bin,
                                'tb_bin': tb_bin,
                                'prob_bin': prob_bin,
                                'total': stats['total'],
                                'hits': stats['hits'],
                                'hit_rate': round(hit_rate, 1),
                                'key': key
                            })

                        # Сортировка
                        reverse = (direction == 'desc')

                        if sort_by == 'prob_bin':
                            # Сортируем по блоку вероятности (числовое значение)
                            def get_prob_value(item):
                                prob_str = item['prob_bin'].replace('%', '')
                                if '-' in prob_str:
                                    low, _ = prob_str.split('-')
                                    return float(low)
                                return 0

                            results_list.sort(key=get_prob_value, reverse=reverse)

                        elif sort_by == 'total':
                            results_list.sort(key=lambda x: x['total'], reverse=reverse)

                        elif sort_by == 'hits':
                            results_list.sort(key=lambda x: x['hits'], reverse=reverse)

                        elif sort_by == 'hit_rate':
                            results_list.sort(key=lambda x: x['hit_rate'], reverse=reverse)

                        # Берем первые 10
                        context['top_results'] = results_list[:limit]
                        context['total_found'] = len(results_list)

            except Exception as e:
                context['error'] = f'Ошибка загрузки данных: {e}'
        else:
            context['error'] = 'Файл с данными не найден'

        return context


@require_POST
@staff_member_required
def bulk_bet_action(request):
    """
    Обработка массовых действий со ставками
    """
    action = request.POST.get('action')
    bet_ids = request.POST.getlist('selected_bets')
    confirm = request.POST.get('confirm')

    if not bet_ids:
        messages.error(request, 'Не выбрано ни одной ставки')
        return redirect('app_bets:records')

    if action == 'delete':
        # Для удаления нужен confirm
        if confirm != 'true':
            # Сохраняем ID в сессии и показываем страницу подтверждения
            request.session['pending_bet_ids'] = bet_ids
            bets = Bet.objects.filter(id__in=bet_ids)
            return render(request, 'app_bets/confirm_bulk_delete.html', {
                'bets': bets,
                'count': bets.count(),
                'total_stake': bets.aggregate(total=Sum('stake'))['total'],
                'total_profit': bets.aggregate(total=Sum('profit'))['total'],
            })
        else:
            # Подтвержденное удаление - берем ID из POST или сессии
            if not bet_ids:
                bet_ids = request.session.pop('pending_bet_ids', [])

            if bet_ids:
                # Получаем все ставки для удаления
                bets_to_delete = Bet.objects.filter(id__in=bet_ids)

                # ВАЖНО: Вызываем delete() для каждой ставки отдельно,
                # чтобы сработал метод delete() модели и обновился банк
                deleted_count = 0
                total_profit_reverted = 0

                for bet in bets_to_delete:
                    profit = bet.profit
                    bet.delete()  # Здесь сработает метод delete() модели
                    if profit:
                        total_profit_reverted += profit
                    deleted_count += 1

                messages.success(
                    request,
                    f'Удалено {deleted_count} ставок. Баланс скорректирован на {abs(total_profit_reverted)} ₽'
                )
            else:
                messages.error(request, 'Не найдены ставки для удаления')

    elif action == 'mark_win':
        bets = Bet.objects.filter(id__in=bet_ids)
        for bet in bets:
            bet.result = Bet.ResultChoices.WIN
            bet.save()  # save() вызовет обновление банка
        messages.success(request, f'Отмечено как выигрыш: {bets.count()} ставок')

    elif action == 'mark_loss':
        bets = Bet.objects.filter(id__in=bet_ids)
        for bet in bets:
            bet.result = Bet.ResultChoices.LOSS
            bet.save()  # save() вызовет обновление банка
        messages.success(request, f'Отмечено как проигрыш: {bets.count()} ставок')

    elif action == 'mark_refund':
        bets = Bet.objects.filter(id__in=bet_ids)
        for bet in bets:
            bet.result = Bet.ResultChoices.REFUND
            bet.save()  # save() вызовет обновление банка
        messages.success(request, f'Отмечено как возврат: {bets.count()} ставок')

    return redirect('app_bets:records')


@staff_member_required
def export_bets_excel(request):
    """
    Экспорт ставок в Excel
    """
    # Получаем отфильтрованный queryset как в BetRecordsView
    queryset = Bet.objects.select_related(
        'home_team', 'away_team', 'league'
    )

    # Применяем те же фильтры, что и в представлении
    # (копируем логику фильтрации из BetRecordsView.get_queryset)
    search_query = request.GET.get('search', '')
    if search_query:
        queryset = queryset.filter(
            Q(home_team__name__icontains=search_query) |
            Q(away_team__name__icontains=search_query) |
            Q(league__name__icontains=search_query)
        )

    date_from = request.GET.get('date_from', '')
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            queryset = queryset.filter(date_placed__date__gte=date_from)
        except ValueError:
            pass

    date_to = request.GET.get('date_to', '')
    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            date_to = datetime.combine(date_to, datetime.max.time())
            queryset = queryset.filter(date_placed__lte=date_to)
        except ValueError:
            pass

    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ставки"

    # Заголовки
    headers = [
        'Дата', 'Время', 'Лига', 'Хозяева', 'Гости',
        'Ставка', 'Кэф', 'Сумма', 'Результат', 'Прибыль',
        'EV %', 'Вер. Пуассона', 'Факт. вер.', 'n матчей'
    ]

    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Данные
    for row_num, bet in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=bet.date_placed.strftime('%d.%m.%Y'))
        ws.cell(row=row_num, column=2, value=bet.match_time)
        ws.cell(row=row_num, column=3, value=str(bet.league))
        ws.cell(row=row_num, column=4, value=bet.home_team.name)
        ws.cell(row=row_num, column=5, value=bet.away_team.name)
        ws.cell(row=row_num, column=6, value=bet.get_recommended_target_display())
        ws.cell(row=row_num, column=7, value=float(bet.recommended_odds))
        ws.cell(row=row_num, column=8, value=float(bet.stake))
        ws.cell(row=row_num, column=9, value=bet.get_result_display())
        ws.cell(row=row_num, column=10, value=float(bet.profit) if bet.profit else 0)
        ws.cell(row=row_num, column=11, value=bet.ev)
        ws.cell(row=row_num, column=12, value=bet.poisson_prob)
        ws.cell(row=row_num, column=13, value=bet.actual_prob)
        ws.cell(row=row_num, column=14, value=bet.n_last_matches)

        # Форматирование чисел
        for col in [7, 8, 10]:
            ws.cell(row=row_num, column=col).number_format = '#,##0.00'
        for col in [11, 12, 13]:
            ws.cell(row=row_num, column=col).number_format = '0.00'

    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width

    # Создаем ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response[
        'Content-Disposition'] = f'attachment; filename=bets_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response

